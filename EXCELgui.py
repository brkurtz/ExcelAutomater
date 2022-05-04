from openpyxl import *
from tkinter import *

__county_data = {

}

def out(outfile) -> None:
    """
    This will output the county data into a different spreadsheet
    :return:
    """
    wb = Workbook()
    ws = wb.active
    # Title new workbook
    ws.title = 'County Data'

    initial = ['State','County','Number of Tracts', 'Pop']
    ws.append(initial)

    for state in __county_data:
        for county in __county_data[state]:
            ws.append([state, county ,__county_data[state][county]['tracts'], __county_data[state][county]['pop']])
    wb.save(outfile)
def parse_sheet(filename) -> None:
    """
    This method will parse the Excel sheet for all states and counties counting the number of tracts and population
    each county has
    :return:
    """

    wb = load_workbook(filename)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        state = ws['B' + str(row)].value
        county = ws['C' + str(row)].value
        pop = ws['D' + str(row)].value

        # Check that the State is in the dictionary
        __county_data.setdefault(state, {})

        # Check that the county exists
        __county_data[state].setdefault(county, {'tracts' : 0,  'pop' : 0})

        # Add to tracts and total pop for each iteration in the loop
        __county_data[state][county]['tracts'] += 1
        __county_data[state][county]['pop'] += int(pop)

class Excelgui:

    def __init__(self, window) -> None:
        """
        Constructor to initialize window for excel application
        """
        # Start
        self.window = window

        # Title
        self.frame_name = Frame(self.window)
        self.label_name = Label(self.frame_name, text='Excel Census Sheet Consolidation', font=('BOLD', 18))
        self.frame_name.pack(side='top', anchor='center')
        self.label_name.pack(side='top', anchor='center')

        # Welcome message
        self.frame_message = Frame(self.window)
        self.label_message = Label(self.frame_message, text="Welcome please enter the name of your sheet with extension:"
                                                        , font=16)
        self.entry_file = Entry(self.frame_message)
        self.frame_message.pack(side='top', anchor='center', pady=10)
        self.label_message.pack(anchor='center',side='left')
        self.entry_file.pack(side='left')

        self.frame_outfile = Frame(self.window)
        self.label_outfile = Label(self.frame_outfile, text='Enter the name of output file with extension:', font= 16)
        self.entry_outfile = Entry(self.frame_outfile)
        self.frame_outfile.pack(side='top', anchor='center', pady= 10)
        self.label_outfile.pack(anchor='center', side='left')
        self.entry_outfile.pack(side='left')

        self.frame_start = Frame(self.window)
        self.button_start = Button(self.frame_start, text='Start', command=self.start)
        self.label_start = Label(self.frame_start, fg='blue')
        self.frame_start.pack(anchor='center')
        self.button_start.pack(anchor='center')
        self.label_start.pack(anchor='center')

    def start(self):
       infile = self.entry_file.get()
       outfile = self.entry_outfile.get()
       parse_sheet(infile)
       out(outfile)
       self.label_start.config(text='Done Compiling Date please check folder with program for sheet.')



